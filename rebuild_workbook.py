"""
rebuild_workbook.py
-------------------
Rebuilds "Quant Strategy(2026-27).xlsx" into a restructured v2 workbook.

Reads the existing workbook (never modifies it) and writes:
    Quant Strategy(2026-27) v2.xlsx

What changes
------------
1. MASTER sheet is the single source of truth for account identity
   (CODE / NAME / ID / STRATEGY / SEGMENT / FUND). Every other sheet pulls
   those fields from it by formula, so they can never drift apart again.
2. Month sheets carry the month total + ROI + days-traded next to the name
   (no scrolling), with every weekday of the month pre-filled as a column.
3. Quarter sheets pull from the LOCAL month sheets with INDEX/MATCH on CODE.
   No external-workbook links, no implicit-intersection VLOOKUP, no source
   range that stops at a hardcoded row.
4. All 12 months + Q1-Q4 + an FY rollup exist from day one.
5. SUMMARY and ISSUES sheets are fully formula-driven.

Excel 2010 compatible: INDEX/MATCH, IFERROR, SUMIF, COUNTIF, LARGE/SMALL only.
No XLOOKUP, no IFNA, no dynamic arrays.

Quarter sheets keep columns A-N in the exact positions update_dashboard.py
reads (CODE=1, NAME=2, ID=3, STRATEGY=4, SEGMENT=5, FUND=6, months 7-12), and
month sheets keep the row-2 "AUG" / "AUG ROI" labels its repair pass matches
on, so the existing dashboard pipeline keeps working unchanged.
"""

import calendar
import sys
from datetime import date

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "Quant Strategy(2026-27).xlsx"
OUT = "Quant Strategy(2026-27) v2.xlsx"

FIRST_ROW = 4       # data starts on row 4 (row 3 = headers, row 2 = labels)
N_ROWS = 120        # pre-provisioned account rows -> 4..123
LAST_ROW = FIRST_ROW + N_ROWS - 1

# (label, sheet name, year, month). The label must match the row-2 text that
# the dashboard's month-detail repair pass looks for.
MONTHS = [
    ("APR", "APR-2026", 2026, 4), ("MAY", "MAY-2026", 2026, 5),
    ("JUNE", "JUNE-2026", 2026, 6), ("JULY", "JULY-2026", 2026, 7),
    ("AUG", "AUG-2026", 2026, 8), ("SEP", "SEP-2026", 2026, 9),
    ("OCT", "OCT-2026", 2026, 10), ("NOV", "NOV-2026", 2026, 11),
    ("DEC", "DEC-2026", 2026, 12), ("JAN", "JAN-2027", 2027, 1),
    ("FEB", "FEB-2027", 2027, 2), ("MAR", "MAR-2027", 2027, 3),
]
MONTH_SHEET = {lab: sh for lab, sh, _, _ in MONTHS}
MONTH_FULL = {
    "APR": "April 2026", "MAY": "May 2026", "JUNE": "June 2026",
    "JULY": "July 2026", "AUG": "August 2026", "SEP": "September 2026",
    "OCT": "October 2026", "NOV": "November 2026", "DEC": "December 2026",
    "JAN": "January 2027", "FEB": "February 2027", "MAR": "March 2027",
}
QUARTERS = [
    ("Q1", "FY-(26-27)Q1", ["APR", "MAY", "JUNE"]),
    ("Q2", "FY-(26-27)Q2", ["JULY", "AUG", "SEP"]),
    ("Q3", "FY-(26-27)Q3", ["OCT", "NOV", "DEC"]),
    ("Q4", "FY-(26-27)Q4", ["JAN", "FEB", "MAR"]),
]
FY_SHEET = "FY-(26-27)"
MASTER = "MASTER"
SUMMARY = "SUMMARY"
ISSUES = "ISSUES"
SCRATCH = "SCRATCH"

# Source sheets, oldest first. Identity conflicts resolve to the most recent
# sheet that carries a non-empty value.
SRC_MONTHS = ["APR-2026", "MAY-2026", "JUNE-2026", "JULY-2026", "AUG-2026"]
SRC_QUARTERS = ["FY-(26-27)Q1", "FY-(26-27)Q2"]

# Month-sheet column plan
C_CODE, C_NAME, C_ID, C_STRAT, C_SEG, C_FUND = 1, 2, 3, 4, 5, 6
C_TOTAL, C_ROI, C_DAYS = 7, 8, 9
C_DAY1 = 10                     # daily date columns start at J

# MASTER column plan
M_CODE, M_NAME, M_DISP, M_ID, M_STRAT, M_SEG, M_FUND, M_STATUS, M_NOTES = range(1, 10)
M_HIST1 = 11                    # fund-history block starts at K (APR..MAR)

# Display names carried over from update_dashboard.py's NAME_OVERRIDES, which
# exist because several accounts share a NAME. Seeded here so the workbook can
# tell them apart too; blank means "use NAME".
DISPLAY_NAMES = {
    "P3224": "Ansh Cash", "P2827": "Bharat ETF", "P2827_2": "Bharat Cash",
    "P3090": "Bharat Straddle", "P2954": "Bharat Munjal_5",
    "P3196": "Bharat Munjal BFO", "P2954_2": "Bharat Munjal S1515",
    "P3079": "Yogesh Cash", "P2777": "Yogesh Kumar", "P2826": "Yogesh ATS",
    "P3039": "Abhishek", "P2967": "Aman", "P2951": "Aman 20-60",
    "P3166": "Aman Dohre", "P3390": "Aman Dohre XTS34", "P3323": "Amit&Kartik",
    "P3323_1": "Amit&Kartik XTS35", "PSW023": "Ansh SWV", "PHPO06": "Ansh HPO",
    "PSW019": "Kartaram SWV097", "SWV0096": "Kartaram SWV096",
    "PHPO07": "Kartaram HPO", "XMR0548": "Kartaram XMR",
    "PA528": "Archana Ma'am", "P3181": "Harshit XTS32", "P3240": "Harshit XTS38",
    "P3311": "Himanshu Pal", "P3109": "Jyoti Prakesh XTS30",
    "P3313": "Jyoti Prakesh ATS", "P3146": "Jinesh Jain", "P2105": "Kamlesh",
    "P3024": "Kartaram_1", "P3202": "Gagandeep", "P3020": "Mahavir Jindal",
    "P3218": "Muhunthan", "P3341": "Maneesh Yadav", "P2971": "Neeraj Garg_1",
    "P2971_2": "Neeraj Garg_2", "P2940": "Prince Cash", "P3347": "Prince FO",
    "P3070": "Prateek_2", "P2999": "Prateek ETF", "P3186": "Prateek NFO",
    "P3135": "Piyush Singhal", "PSW024": "Piyush SWV", "P3168": "Prabjot Singh",
    "P3342": "Raghav Tuli", "P3082": "Ramakar Jha", "P3048": "Shahid",
    "P3208": "Sajal Sharma ATS", "P3335": "Sajal Sharma_2",
    "P3119": "Sajal Sharma_3", "P3297": "Sudeep", "P3360": "Vidhya Sagar",
    "P3385": "Vikas Gupta XTS13", "P3386": "Vikas Gupta XTS15",
    "P3334": "Varun Tondan XTS10", "P3110": "Varun Tondan", "P3113": "Vaishali",
    "P2817": "DK Sir_M", "P2792": "DK Sir_T", "P3013": "Deepesh_1",
    "P3112": "Deepesh_2", "P3361": "Nishaanth",
}

# ---------------------------------------------------------------- styling ----
FMT_PNL = '#,##,##0;[Red](#,##,##0);"-"'
FMT_ROI = '0.00"%";[Red](0.00"%");"-"'
FMT_FUND = '0.00'
FMT_INT = '0;;"-"'
FMT_DATE = 'dd-mmm'

NAVY, SLATE, LIGHT, BAND = "1F3864", "2F5496", "D9E2F3", "F7F9FD"
GREY, AMBER, REDBG, GRNBG = "808080", "FFF2CC", "FFC7CE", "C6EFCE"

F_TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
F_SUB = Font(name="Calibri", size=9, italic=True, color=GREY)
F_HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_SECT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_LBL = Font(name="Calibri", size=10, bold=True, color=NAVY)
F_BODY = Font(name="Calibri", size=10)
F_KPI = Font(name="Calibri", size=13, bold=True, color=NAVY)

FILL_TITLE = PatternFill("solid", fgColor=NAVY)
FILL_HDR = PatternFill("solid", fgColor=SLATE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_AMBER = PatternFill("solid", fgColor=AMBER)

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def norm(v):
    return "" if v is None else str(v).strip()


def weekdays(year, month):
    n = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, n + 1)
            if date(year, month, d).weekday() < 5]


def title_bar(ws, text, sub, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, text)
    c.font, c.fill, c.alignment = F_TITLE, FILL_TITLE, LEFT
    ws.row_dimensions[1].height = 22
    if sub:
        ws.cell(2, 1, sub).font = F_SUB


def header_cell(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.font, c.fill, c.alignment, c.border = F_HDR, FILL_HDR, CTR, BOX
    return c


# ============================================================ 1. READ SOURCE ==
def read_source():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    accounts = {}        # code -> identity dict
    order = []           # code order, first-seen
    daily = {}           # (code, date) -> value
    fund_by_month = {}   # (code, label) -> fund recorded on that month sheet
    conflicts = []       # [severity, area, detail]
    placeholders = []    # rows that had no CODE

    def touch(code):
        if code not in accounts:
            accounts[code] = dict(code=code, name="", id="", strat="", seg="",
                                  fund=None, seen=[])
            order.append(code)
        return accounts[code]

    variants = {}   # code -> field -> {value seen}

    def absorb(a, label, name, _id, strat, seg, fund, is_month):
        for field, val in (("name", name), ("id", _id),
                           ("strat", strat), ("seg", seg)):
            if not val:
                continue
            variants.setdefault(a["code"], {}).setdefault(field, set()).add(val)
            if is_month or not a[field]:
                a[field] = val
        if fund is not None and (is_month or a["fund"] is None):
            a["fund"] = fund
        a["seen"].append(label)

    for sheet in SRC_MONTHS:
        ws = wb[sheet]
        label = sheet.split("-")[0]
        datecols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(2, c).value
            if hasattr(v, "year"):
                datecols[c] = v.date() if hasattr(v, "date") else v
        for r in range(FIRST_ROW, ws.max_row + 1):
            code = norm(ws.cell(r, 1).value)
            name = norm(ws.cell(r, 2).value)
            _id = norm(ws.cell(r, 3).value)
            if not code:
                if not (name or _id):
                    continue
                code = "X_" + (_id or name).upper().replace(" ", "")[:8]
                placeholders.append((sheet, r, name, _id, code))
            fund = ws.cell(r, 6).value
            fundf = float(fund) if isinstance(fund, (int, float)) else None
            a = touch(code)
            absorb(a, sheet, name, _id, norm(ws.cell(r, 4).value),
                   norm(ws.cell(r, 5).value), fundf, True)
            # Record the fund this month actually used, blank included. A month
            # that carried no fund produced an ROI of 0 in the old workbook
            # (IFERROR swallowed the divide), so record 0 to reproduce it
            # rather than silently back-filling today's fund into the past.
            fund_by_month[(code, label)] = fundf if fundf is not None else 0.0
            for c, d in datecols.items():
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v != 0:
                    daily[(code, d)] = v

    for sheet in SRC_QUARTERS:
        ws = wb[sheet]
        for r in range(FIRST_ROW, ws.max_row + 1):
            code = norm(ws.cell(r, 1).value)
            name = norm(ws.cell(r, 2).value)
            _id = norm(ws.cell(r, 3).value)
            if not code:
                if not (name or _id):
                    continue
                code = "X_" + (_id or name).upper().replace(" ", "")[:8]
            fund = ws.cell(r, 6).value
            fundf = float(fund) if isinstance(fund, (int, float)) else None
            absorb(touch(code), sheet, name, _id, norm(ws.cell(r, 4).value),
                   norm(ws.cell(r, 5).value), fundf, False)

    # An account with no row at all on a past month sheet carried no capital
    # that month. Pin it to 0 so SUMMARY does not count today's fund against a
    # month the account was not trading in.
    for lab in (s.split("-")[0] for s in SRC_MONTHS):
        for code in order:
            fund_by_month.setdefault((code, lab), 0.0)

    # one row per account per field that drifted, not one per sheet it drifted on
    FIELD_LABEL = {"name": "NAME", "id": "ID", "strat": "STRATEGY", "seg": "SEGMENT"}
    for code in order:
        for field, seen in sorted(variants.get(code, {}).items()):
            if len({v.lower() for v in seen}) < 2:
                continue
            kept = accounts[code][field]
            others = sorted(v for v in seen if v.lower() != kept.lower())
            conflicts.append([
                "Check", FIELD_LABEL[field] + " was inconsistent",
                "%s was written %d different ways across sheets: %s. MASTER now "
                "uses '%s' everywhere -- change it there if that is the wrong one."
                % (code, len(seen), ", ".join("'%s'" % v for v in others + [kept]),
                   kept)])

    # A month with no fund only matters if the account actually traded that
    # month -- otherwise it simply had not opened yet.
    traded = set()
    for (code, d) in daily:
        for lab, _, year, month in MONTHS:
            if d.year == year and d.month == month:
                traded.add((code, lab))
                break
    for (code, lab), v in sorted(fund_by_month.items()):
        cur = accounts[code]["fund"]
        if v == 0 and cur and (code, lab) in traded:
            conflicts.append([
                "Check", "Traded with no fund recorded",
                "%s has %s P&L but no fund on the %s sheet, so its %s ROI stays "
                "0.00%%. Current fund is %.2f Cr -- put the right figure in "
                "MASTER's %s fund-history cell to get a real ROI."
                % (code, lab, lab, lab, cur, lab)])

    return accounts, order, daily, fund_by_month, conflicts, placeholders


# ================================================================ 2. MASTER ==
def build_master(wb, accounts, order, fund_by_month):
    ws = wb.create_sheet(MASTER)
    title_bar(ws, "MASTER  -  account registry",
              "The single source of truth. Add or edit an account HERE ONLY; "
              "every month, quarter and FY sheet follows automatically. "
              "Rows 4-%d are wired -- fill the next empty row to add an account."
              % LAST_ROW, 22)

    heads = ["CODE", "NAME", "DISPLAY NAME", "ID", "STRATEGY", "SEGMENT",
             "FUND (CR)", "STATUS", "NOTES"]
    for i, h in enumerate(heads, start=1):
        header_cell(ws, 3, i, h)
    ws.cell(2, M_HIST1, "FUND HISTORY (CR)  -  leave blank to use FUND (CR)").font = F_LBL
    for i, (lab, _, _, _) in enumerate(MONTHS):
        header_cell(ws, 3, M_HIST1 + i, lab)

    segments = sorted({a["seg"] for a in accounts.values() if a["seg"]})

    for i, code in enumerate(order):
        a = accounts[code]
        r = FIRST_ROW + i
        vals = [code, a["name"], DISPLAY_NAMES.get(code, ""), a["id"],
                a["strat"], a["seg"],
                a["fund"] if a["fund"] is not None else 0,
                "Active", ""]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(r, ci, v)
            c.font, c.border = F_BODY, BOX
            c.alignment = CTR if ci in (M_CODE, M_SEG, M_STATUS) else LEFT
        ws.cell(r, M_FUND).number_format = FMT_FUND
        # fund history: only where a month recorded something other than current
        for mi, (lab, _, _, _) in enumerate(MONTHS):
            v = fund_by_month.get((code, lab))
            c = ws.cell(r, M_HIST1 + mi)
            c.font, c.border, c.number_format = F_BODY, BOX, FMT_FUND
            if v is not None and a["fund"] is not None and abs(v - a["fund"]) > 1e-9:
                c.value = v
                c.fill = FILL_AMBER

    # empty spare rows keep their formatting and validation
    for r in range(FIRST_ROW + len(order), LAST_ROW + 1):
        for ci in range(1, M_NOTES + 1):
            c = ws.cell(r, ci)
            c.font, c.border = F_BODY, BOX
            c.alignment = CTR if ci in (M_CODE, M_SEG, M_STATUS) else LEFT
        ws.cell(r, M_FUND).number_format = FMT_FUND
        for mi in range(len(MONTHS)):
            c = ws.cell(r, M_HIST1 + mi)
            c.font, c.border, c.number_format = F_BODY, BOX, FMT_FUND

    dv_seg = DataValidation(type="list", allow_blank=True,
                            formula1='"%s"' % ",".join(segments)[:250],
                            showErrorMessage=False)
    ws.add_data_validation(dv_seg)
    dv_seg.add("%s%d:%s%d" % (L(M_SEG), FIRST_ROW, L(M_SEG), LAST_ROW))

    dv_st = DataValidation(type="list", allow_blank=True,
                           formula1='"Active,Paused,Closed"', showErrorMessage=False)
    ws.add_data_validation(dv_st)
    dv_st.add("%s%d:%s%d" % (L(M_STATUS), FIRST_ROW, L(M_STATUS), LAST_ROW))

    dv_dup = DataValidation(
        type="custom",
        formula1="COUNTIF($A$%d:$A$%d,A%d)<=1" % (FIRST_ROW, LAST_ROW, FIRST_ROW),
        allow_blank=True, showErrorMessage=True,
        errorTitle="Duplicate CODE",
        error="That CODE already exists in MASTER. Every account needs a unique code.")
    ws.add_data_validation(dv_dup)
    dv_dup.add("A%d:A%d" % (FIRST_ROW, LAST_ROW))

    # highlight a duplicated code even if one was pasted in past the validation
    ws.conditional_formatting.add(
        "A%d:A%d" % (FIRST_ROW, LAST_ROW),
        FormulaRule(formula=["AND($A%d<>\"\",COUNTIF($A$%d:$A$%d,$A%d)>1)"
                             % (FIRST_ROW, FIRST_ROW, LAST_ROW, FIRST_ROW)],
                    fill=PatternFill("solid", fgColor=REDBG)))

    widths = {1: 11, 2: 22, 3: 22, 4: 12, 5: 18, 6: 11, 7: 11, 8: 9, 9: 26}
    for c, w in widths.items():
        ws.column_dimensions[L(c)].width = w
    for i in range(len(MONTHS)):
        ws.column_dimensions[L(M_HIST1 + i)].width = 7
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = "A3:I%d" % LAST_ROW
    return ws


# ========================================================== 3. MONTH SHEETS ==
def identity_formulas(ws, r):
    """A-F pulled from MASTER by row position, blank when MASTER row is empty."""
    ws.cell(r, C_CODE, "=IF(%s!$A%d=\"\",\"\",%s!$A%d)" % (MASTER, r, MASTER, r))
    ws.cell(r, C_NAME, "=IF($A%d=\"\",\"\",IF(%s!$C%d=\"\",%s!$B%d,%s!$C%d))"
            % (r, MASTER, r, MASTER, r, MASTER, r))
    ws.cell(r, C_ID, "=IF($A%d=\"\",\"\",%s!$D%d)" % (r, MASTER, r))
    ws.cell(r, C_STRAT, "=IF($A%d=\"\",\"\",%s!$E%d)" % (r, MASTER, r))
    ws.cell(r, C_SEG, "=IF($A%d=\"\",\"\",%s!$F%d)" % (r, MASTER, r))


def build_month(wb, label, sheet, year, month, daily, order, accounts):
    ws = wb.create_sheet(sheet)
    days = weekdays(year, month)
    last_day_col = C_DAY1 + len(days) - 1
    width = last_day_col

    title_bar(ws, "%s  -  daily P&L" % MONTH_FULL[label].upper(),
              "Type only in the dated columns. Name, strategy, segment and fund "
              "come from MASTER; the total, ROI and days columns are formulas.",
              width)

    ws.cell(2, C_TOTAL, label).font = F_LBL          # dashboard matches on these
    ws.cell(2, C_ROI, label + " ROI").font = F_LBL
    for i, h in enumerate(["CODE", "NAME", "ID", "STRATEGY", "SEGMENT",
                           "FUND (CR)", label + " TOTAL", "ROI %", "DAYS"], start=1):
        header_cell(ws, 3, i, h)

    for i, d in enumerate(days):
        col = C_DAY1 + i
        c = ws.cell(2, col, d)
        c.number_format, c.font, c.alignment = FMT_DATE, F_LBL, CTR
        c.border = BOX
        h = ws.cell(3, col, ["MON", "TUE", "WED", "THU", "FRI"][d.weekday()])
        h.font, h.fill, h.alignment, h.border = F_HDR, FILL_HDR, CTR, BOX

    mi = [m[0] for m in MONTHS].index(label)
    hist = L(M_HIST1 + mi)
    first_day, last_day = L(C_DAY1), L(last_day_col)

    for r in range(FIRST_ROW, LAST_ROW + 1):
        identity_formulas(ws, r)
        # fund: this month's history value if present, else the current fund
        ws.cell(r, C_FUND,
                "=IF($A%d=\"\",\"\",IF(%s!$%s%d=\"\",%s!$G%d,%s!$%s%d))"
                % (r, MASTER, hist, r, MASTER, r, MASTER, hist, r))
        ws.cell(r, C_TOTAL,
                "=IF($A%d=\"\",\"\",SUM(%s%d:%s%d))" % (r, first_day, r, last_day, r))
        ws.cell(r, C_ROI,
                "=IF($A%d=\"\",\"\",IFERROR($G%d/($F%d*10000000)*100,0))" % (r, r, r))
        ws.cell(r, C_DAYS,
                "=IF($A%d=\"\",\"\",COUNTIF(%s%d:%s%d,\"<>0\"))"
                % (r, first_day, r, last_day, r))

        band = FILL_BAND if r % 2 == 0 else None
        for c in range(1, width + 1):
            cell = ws.cell(r, c)
            cell.font, cell.border = F_BODY, BOX
            if band:
                cell.fill = band
            if c in (C_CODE, C_ID, C_SEG):
                cell.alignment = CTR
            elif c in (C_NAME, C_STRAT):
                cell.alignment = LEFT
            else:
                cell.alignment = CTR
        ws.cell(r, C_FUND).number_format = FMT_FUND
        ws.cell(r, C_TOTAL).number_format = FMT_PNL
        ws.cell(r, C_ROI).number_format = FMT_ROI
        ws.cell(r, C_DAYS).number_format = FMT_INT
        for c in range(C_DAY1, width + 1):
            ws.cell(r, c).number_format = FMT_PNL

    # carry the existing numbers over
    filled = 0
    for i, code in enumerate(order):
        r = FIRST_ROW + i
        for j, d in enumerate(days):
            v = daily.get((code, d))
            if v is not None:
                ws.cell(r, C_DAY1 + j, v)
                filled += 1

    total_row = LAST_ROW + 2
    ws.cell(total_row, C_NAME, "TOTAL").font = F_LBL
    tc = ws.cell(total_row, C_TOTAL, "=SUM(%s%d:%s%d)" % (L(C_TOTAL), FIRST_ROW,
                                                          L(C_TOTAL), LAST_ROW))
    tc.number_format, tc.font, tc.fill, tc.border = FMT_PNL, F_LBL, FILL_LIGHT, BOX
    fc = ws.cell(total_row, C_FUND, "=SUM(%s%d:%s%d)" % (L(C_FUND), FIRST_ROW,
                                                         L(C_FUND), LAST_ROW))
    fc.number_format, fc.font, fc.fill, fc.border = FMT_FUND, F_LBL, FILL_LIGHT, BOX
    rc = ws.cell(total_row, C_ROI,
                 "=IFERROR($G%d/($F%d*10000000)*100,0)" % (total_row, total_row))
    rc.number_format, rc.font, rc.fill, rc.border = FMT_ROI, F_LBL, FILL_LIGHT, BOX
    for c in range(C_DAY1, width + 1):
        d = ws.cell(total_row, c,
                    "=SUM(%s%d:%s%d)" % (L(c), FIRST_ROW, L(c), LAST_ROW))
        d.number_format, d.font, d.fill, d.border = FMT_PNL, F_LBL, FILL_LIGHT, BOX

    rng = "%s%d:%s%d" % (L(C_TOTAL), FIRST_ROW, L(C_TOTAL), LAST_ROW)
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=REDBG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=GRNBG)))

    for c, w in {1: 11, 2: 22, 3: 11, 4: 17, 5: 10, 6: 10, 7: 14, 8: 10, 9: 7}.items():
        ws.column_dimensions[L(c)].width = w
    for c in range(C_DAY1, width + 1):
        ws.column_dimensions[L(c)].width = 11
    ws.freeze_panes = "%s4" % L(C_DAY1)
    ws.auto_filter.ref = "A3:%s%d" % (L(C_DAYS), LAST_ROW)
    return filled


# ======================================================== 4. QUARTER SHEETS ==
def build_quarter(wb, qkey, sheet, labels):
    ws = wb.create_sheet(sheet)
    width = 16
    title_bar(ws, "%s  -  %s" % (qkey, " / ".join(MONTH_FULL[l] for l in labels)),
              "Every number is pulled from the month sheets by CODE. Nothing on "
              "this sheet is typed by hand.", width)

    pairs = [(7, 8), (9, 10), (11, 12)]
    for lab, (cp, cr) in zip(labels, pairs):
        ws.cell(2, cp, lab).font = F_LBL
        ws.cell(2, cr, lab + " ROI").font = F_LBL
    ws.cell(2, 13, qkey.replace("Q", "Q ")).font = F_LBL
    ws.cell(2, 14, "ROI " + qkey).font = F_LBL

    heads = ["CODE", "NAME", "ID", "STRATEGY", "SEGMENT", "FUND (CR)"]
    for lab in labels:
        heads += [lab, lab + " ROI"]
    heads += [qkey + " TOTAL", "ROI " + qkey, "", "DATA CHECK"]
    for i, h in enumerate(heads, start=1):
        if h:
            header_cell(ws, 3, i, h)

    for r in range(FIRST_ROW, LAST_ROW + 1):
        identity_formulas(ws, r)
        ws.cell(r, C_FUND, "=IF($A%d=\"\",\"\",%s!$G%d)" % (r, MASTER, r))
        for lab, (cp, cr) in zip(labels, pairs):
            sh = MONTH_SHEET[lab]
            for col, src in ((cp, "G"), (cr, "H")):
                ws.cell(r, col,
                        "=IF($A%d=\"\",\"\",IFERROR(INDEX('%s'!$%s$%d:$%s$%d,"
                        "MATCH($A%d,'%s'!$A$%d:$A$%d,0)),0))"
                        % (r, sh, src, FIRST_ROW, src, LAST_ROW,
                           r, sh, FIRST_ROW, LAST_ROW))
        ws.cell(r, 13, "=IF($A%d=\"\",\"\",$G%d+$I%d+$K%d)" % (r, r, r, r))
        ws.cell(r, 14, "=IF($A%d=\"\",\"\",$H%d+$J%d+$L%d)" % (r, r, r, r))
        checks = "+".join(
            "COUNTIF('%s'!$A$%d:$A$%d,$A%d)" % (MONTH_SHEET[l], FIRST_ROW, LAST_ROW, r)
            for l in labels)
        ws.cell(r, 16, "=IF($A%d=\"\",\"\",IF(%s=3,\"OK\",\"NOT IN ALL 3 MONTHS\"))"
                % (r, checks))

        band = FILL_BAND if r % 2 == 0 else None
        for c in list(range(1, 15)) + [16]:
            cell = ws.cell(r, c)
            cell.font, cell.border = F_BODY, BOX
            cell.alignment = LEFT if c in (C_NAME, C_STRAT) else CTR
            if band:
                cell.fill = band
        ws.cell(r, C_FUND).number_format = FMT_FUND
        for cp, cr in pairs:
            ws.cell(r, cp).number_format = FMT_PNL
            ws.cell(r, cr).number_format = FMT_ROI
        ws.cell(r, 13).number_format = FMT_PNL
        ws.cell(r, 14).number_format = FMT_ROI

    total_row = LAST_ROW + 2
    ws.cell(total_row, C_NAME, "TOTAL").font = F_LBL
    for c in [C_FUND, 7, 9, 11, 13]:
        t = ws.cell(total_row, c, "=SUM(%s%d:%s%d)" % (L(c), FIRST_ROW, L(c), LAST_ROW))
        t.number_format = FMT_FUND if c == C_FUND else FMT_PNL
        t.font, t.fill, t.border = F_LBL, FILL_LIGHT, BOX
    for c, src in ((8, 7), (10, 9), (12, 11), (14, 13)):
        t = ws.cell(total_row, c, "=IFERROR(%s%d/($F%d*10000000)*100,0)"
                    % (L(src), total_row, total_row))
        t.number_format, t.font, t.fill, t.border = FMT_ROI, F_LBL, FILL_LIGHT, BOX

    # reconciliation tripwire: quarter column vs the month sheet's own total
    rec = LAST_ROW + 3
    ws.cell(rec, C_NAME, "RECONCILE vs month sheet").font = F_LBL
    for lab, (cp, _) in zip(labels, pairs):
        sh = MONTH_SHEET[lab]
        t = ws.cell(rec, cp, "=%s%d-SUM('%s'!$G$%d:$G$%d)"
                    % (L(cp), total_row, sh, FIRST_ROW, LAST_ROW))
        t.number_format, t.font, t.border = FMT_PNL, F_LBL, BOX
        ws.conditional_formatting.add(
            "%s%d" % (L(cp), rec),
            CellIsRule(operator="notEqual", formula=["0"],
                       fill=PatternFill("solid", fgColor=REDBG)))
    ws.cell(rec, 16, "must be zero -- any other value means a row was lost").font = F_SUB

    rng = "M%d:M%d" % (FIRST_ROW, LAST_ROW)
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=REDBG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=GRNBG)))
    ws.conditional_formatting.add(
        "P%d:P%d" % (FIRST_ROW, LAST_ROW),
        CellIsRule(operator="equal", formula=['"NOT IN ALL 3 MONTHS"'],
                   fill=PatternFill("solid", fgColor=REDBG)))

    for c, w in {1: 11, 2: 22, 3: 11, 4: 17, 5: 10, 6: 10, 7: 14, 8: 10, 9: 14,
                 10: 10, 11: 14, 12: 10, 13: 15, 14: 10, 15: 3, 16: 22}.items():
        ws.column_dimensions[L(c)].width = w
    ws.freeze_panes = "G4"
    ws.auto_filter.ref = "A3:N%d" % LAST_ROW


# ============================================================= 5. FY ROLLUP ==
def build_fy(wb):
    ws = wb.create_sheet(FY_SHEET)
    width = 18
    title_bar(ws, "FY 2026-27  -  full year by account",
              "Q1-Q4 pulled from the quarter sheets by CODE. Column R is a "
              "hidden ranking helper used by SUMMARY.", width)

    for i, (qkey, _, _) in enumerate(QUARTERS):
        ws.cell(2, 7 + i * 2, qkey).font = F_LBL
        ws.cell(2, 8 + i * 2, "ROI " + qkey).font = F_LBL
    ws.cell(2, 15, "FY 2026-27").font = F_LBL
    ws.cell(2, 16, "ROI FY").font = F_LBL

    heads = ["CODE", "NAME", "ID", "STRATEGY", "SEGMENT", "FUND (CR)"]
    for qkey, _, _ in QUARTERS:
        heads += [qkey, "ROI " + qkey]
    heads += ["FY TOTAL", "ROI FY", "", "RANK HELPER"]
    for i, h in enumerate(heads, start=1):
        if h:
            header_cell(ws, 3, i, h)

    for r in range(FIRST_ROW, LAST_ROW + 1):
        identity_formulas(ws, r)
        ws.cell(r, C_FUND, "=IF($A%d=\"\",\"\",%s!$G%d)" % (r, MASTER, r))
        for i, (_, sh, _) in enumerate(QUARTERS):
            for off, src in ((0, "M"), (1, "N")):
                ws.cell(r, 7 + i * 2 + off,
                        "=IF($A%d=\"\",\"\",IFERROR(INDEX('%s'!$%s$%d:$%s$%d,"
                        "MATCH($A%d,'%s'!$A$%d:$A$%d,0)),0))"
                        % (r, sh, src, FIRST_ROW, src, LAST_ROW,
                           r, sh, FIRST_ROW, LAST_ROW))
        ws.cell(r, 15, "=IF($A%d=\"\",\"\",$G%d+$I%d+$K%d+$M%d)" % (r, r, r, r, r))
        ws.cell(r, 16, "=IF($A%d=\"\",\"\",$H%d+$J%d+$L%d+$N%d)" % (r, r, r, r, r))
        ws.cell(r, 18, "=IF($A%d=\"\",\"\",$P%d-ROW()/1000000)" % (r, r))

        band = FILL_BAND if r % 2 == 0 else None
        for c in range(1, 17):
            cell = ws.cell(r, c)
            cell.font, cell.border = F_BODY, BOX
            cell.alignment = LEFT if c in (C_NAME, C_STRAT) else CTR
            if band:
                cell.fill = band
        ws.cell(r, C_FUND).number_format = FMT_FUND
        for i in range(4):
            ws.cell(r, 7 + i * 2).number_format = FMT_PNL
            ws.cell(r, 8 + i * 2).number_format = FMT_ROI
        ws.cell(r, 15).number_format = FMT_PNL
        ws.cell(r, 16).number_format = FMT_ROI

    total_row = LAST_ROW + 2
    ws.cell(total_row, C_NAME, "TOTAL").font = F_LBL
    for c in [C_FUND, 7, 9, 11, 13, 15]:
        t = ws.cell(total_row, c, "=SUM(%s%d:%s%d)" % (L(c), FIRST_ROW, L(c), LAST_ROW))
        t.number_format = FMT_FUND if c == C_FUND else FMT_PNL
        t.font, t.fill, t.border = F_LBL, FILL_LIGHT, BOX
    for c, src in ((8, 7), (10, 9), (12, 11), (14, 13), (16, 15)):
        t = ws.cell(total_row, c, "=IFERROR(%s%d/($F%d*10000000)*100,0)"
                    % (L(src), total_row, total_row))
        t.number_format, t.font, t.fill, t.border = FMT_ROI, F_LBL, FILL_LIGHT, BOX

    rng = "O%d:O%d" % (FIRST_ROW, LAST_ROW)
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=REDBG)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=GRNBG)))

    for c, w in {1: 11, 2: 22, 3: 11, 4: 17, 5: 10, 6: 10}.items():
        ws.column_dimensions[L(c)].width = w
    for c in range(7, 17):
        ws.column_dimensions[L(c)].width = 14 if c % 2 else 10
    ws.column_dimensions["Q"].width = 3
    ws.column_dimensions["R"].hidden = True
    ws.freeze_panes = "G4"
    ws.auto_filter.ref = "A3:P%d" % LAST_ROW


# =============================================================== 6. SUMMARY ==
def build_summary(wb, accounts):
    ws = wb.create_sheet(SUMMARY)
    title_bar(ws, "QUANT STRATEGY  -  FY 2026-27 SUMMARY",
              "Every figure is a live formula. Nothing here is typed by hand.", 8)

    def section(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row, 1, text)
        c.font, c.fill, c.alignment = F_SECT, FILL_HDR, LEFT
        return row + 1

    def table_head(row, labels, start=1):
        for i, h in enumerate(labels):
            header_cell(ws, row, start + i, h)
        return row + 1

    fy = "'%s'" % FY_SHEET
    fy_pnl = "%s!$O$%d:$O$%d" % (fy, FIRST_ROW, LAST_ROW)
    fy_roi = "%s!$P$%d:$P$%d" % (fy, FIRST_ROW, LAST_ROW)
    fy_rank = "%s!$R$%d:$R$%d" % (fy, FIRST_ROW, LAST_ROW)
    fy_name = "%s!$B$%d:$B$%d" % (fy, FIRST_ROW, LAST_ROW)
    fy_seg = "%s!$E$%d:$E$%d" % (fy, FIRST_ROW, LAST_ROW)
    fy_strat = "%s!$D$%d:$D$%d" % (fy, FIRST_ROW, LAST_ROW)
    fy_fund = "%s!$F$%d:$F$%d" % (fy, FIRST_ROW, LAST_ROW)
    m_fund = "%s!$G$%d:$G$%d" % (MASTER, FIRST_ROW, LAST_ROW)

    r = section(4, "HEADLINE")
    kpis = [
        ("Net P&L (FY to date)", "=SUM(%s)" % fy_pnl, FMT_PNL),
        ("Fund deployed (Cr)", "=SUM(%s)" % m_fund, FMT_FUND),
        ("Return on fund", "=IFERROR($B$%d/($B$%d*10000000)*100,0)" % (r, r + 1), FMT_ROI),
        ("Accounts on book", "=COUNTA(%s!$A$%d:$A$%d)" % (MASTER, FIRST_ROW, LAST_ROW), "0"),
        ("Accounts in profit", "=COUNTIF(%s,\">0\")" % fy_pnl, "0"),
        ("Accounts in loss", "=COUNTIF(%s,\"<0\")" % fy_pnl, "0"),
        ("Win rate", "=IFERROR($B$%d/($B$%d+$B$%d)*100,0)" % (r + 4, r + 4, r + 5), FMT_ROI),
    ]
    for i, (lab, f, fmt) in enumerate(kpis):
        ws.cell(r + i, 1, lab).font = F_LBL
        c = ws.cell(r + i, 2, f)
        c.number_format, c.font, c.fill, c.border = fmt, F_KPI, FILL_LIGHT, BOX
    r += len(kpis) + 1

    r = section(r, "BY QUARTER")
    r = table_head(r, ["Quarter", "Net P&L", "ROI %", "Fund (Cr)", "Accounts traded"])
    for i, (qkey, sh, labels) in enumerate(QUARTERS):
        pcol = L(7 + i * 2)
        ws.cell(r + i, 1, "%s  (%s)" % (qkey, ", ".join(labels))).font = F_BODY
        ws.cell(r + i, 2, "=SUM(%s!$%s$%d:$%s$%d)"
                % (fy, pcol, FIRST_ROW, pcol, LAST_ROW)).number_format = FMT_PNL
        ws.cell(r + i, 3, "=IFERROR($B%d/($D%d*10000000)*100,0)"
                % (r + i, r + i)).number_format = FMT_ROI
        # average capital across the quarter's three months, so the ROI is
        # measured against what was actually deployed, not today's book
        ws.cell(r + i, 4, "=(%s)/3" % "+".join(
            "SUM('%s'!$F$%d:$F$%d)" % (MONTH_SHEET[l], FIRST_ROW, LAST_ROW)
            for l in labels)).number_format = FMT_FUND
        ws.cell(r + i, 5, "=COUNTIF(%s!$%s$%d:$%s$%d,\"<>0\")-COUNTBLANK(%s!$%s$%d:$%s$%d)"
                % (fy, pcol, FIRST_ROW, pcol, LAST_ROW,
                   fy, pcol, FIRST_ROW, pcol, LAST_ROW)).number_format = "0"
        for c in range(1, 6):
            ws.cell(r + i, c).border = BOX
    r += len(QUARTERS) + 1

    r = section(r, "BY MONTH")
    r = table_head(r, ["Month", "Net P&L", "ROI %", "Fund (Cr)", "Days traded"])
    for i, (lab, sh, _, _) in enumerate(MONTHS):
        ws.cell(r + i, 1, MONTH_FULL[lab]).font = F_BODY
        ws.cell(r + i, 2, "=SUM('%s'!$G$%d:$G$%d)"
                % (sh, FIRST_ROW, LAST_ROW)).number_format = FMT_PNL
        ws.cell(r + i, 3, "=IFERROR($B%d/($D%d*10000000)*100,0)"
                % (r + i, r + i)).number_format = FMT_ROI
        ws.cell(r + i, 4, "=SUM('%s'!$F$%d:$F$%d)"
                % (sh, FIRST_ROW, LAST_ROW)).number_format = FMT_FUND
        ws.cell(r + i, 5, "=MAX('%s'!$I$%d:$I$%d)"
                % (sh, FIRST_ROW, LAST_ROW)).number_format = "0"
        for c in range(1, 6):
            ws.cell(r + i, c).border = BOX
    r += len(MONTHS) + 1

    segs = sorted({a["seg"] for a in accounts.values() if a["seg"]})
    r = section(r, "BY SEGMENT")
    r = table_head(r, ["Segment", "Net P&L (FY)", "Fund (Cr)", "ROI %", "Accounts"])
    for i, s in enumerate(segs):
        ws.cell(r + i, 1, s).font = F_BODY
        ws.cell(r + i, 2, "=SUMIF(%s,$A%d,%s)" % (fy_seg, r + i, fy_pnl)).number_format = FMT_PNL
        ws.cell(r + i, 3, "=SUMIF(%s,$A%d,%s)" % (fy_seg, r + i, fy_fund)).number_format = FMT_FUND
        ws.cell(r + i, 4, "=IFERROR($B%d/($C%d*10000000)*100,0)" % (r + i, r + i)).number_format = FMT_ROI
        ws.cell(r + i, 5, "=COUNTIF(%s,$A%d)" % (fy_seg, r + i)).number_format = "0"
        for c in range(1, 6):
            ws.cell(r + i, c).border = BOX
    r += len(segs) + 1

    strats = sorted({a["strat"] for a in accounts.values() if a["strat"]},
                    key=lambda s: s.upper())
    r = section(r, "BY STRATEGY")
    r = table_head(r, ["Strategy", "Net P&L (FY)", "Fund (Cr)", "ROI %", "Accounts"])
    for i, s in enumerate(strats):
        ws.cell(r + i, 1, s).font = F_BODY
        ws.cell(r + i, 2, "=SUMIF(%s,$A%d,%s)" % (fy_strat, r + i, fy_pnl)).number_format = FMT_PNL
        ws.cell(r + i, 3, "=SUMIF(%s,$A%d,%s)" % (fy_strat, r + i, fy_fund)).number_format = FMT_FUND
        ws.cell(r + i, 4, "=IFERROR($B%d/($C%d*10000000)*100,0)" % (r + i, r + i)).number_format = FMT_ROI
        ws.cell(r + i, 5, "=COUNTIF(%s,$A%d)" % (fy_strat, r + i)).number_format = "0"
        for c in range(1, 6):
            ws.cell(r + i, c).border = BOX
    r += len(strats) + 1

    for heading, fn in (("TOP 5 BY FY ROI", "LARGE"), ("BOTTOM 5 BY FY ROI", "SMALL")):
        r = section(r, heading)
        r = table_head(r, ["#", "Account", "FY P&L", "ROI %", "Fund (Cr)"])
        for k in range(1, 6):
            row = r + k - 1
            pick = "%s(%s,%d)" % (fn, fy_rank, k)
            ws.cell(row, 1, k).font = F_BODY
            for col, src in ((2, fy_name), (3, fy_pnl), (4, fy_roi), (5, fy_fund)):
                ws.cell(row, col, "=IFERROR(INDEX(%s,MATCH(%s,%s,0)),\"\")"
                        % (src, pick, fy_rank))
            ws.cell(row, 3).number_format = FMT_PNL
            ws.cell(row, 4).number_format = FMT_ROI
            ws.cell(row, 5).number_format = FMT_FUND
            for c in range(1, 6):
                ws.cell(row, c).border = BOX
        r += 6

    for c, w in {1: 34, 2: 18, 3: 14, 4: 14, 5: 16}.items():
        ws.column_dimensions[L(c)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ================================================================ 7. ISSUES ==
def build_issues(wb, conflicts, placeholders, accounts):
    ws = wb.create_sheet(ISSUES)
    title_bar(ws, "ISSUES  -  data health",
              "The top block is live: it recalculates every time the workbook "
              "opens. The bottom block is what the rebuild found in the old file.", 6)

    r = 4
    c = ws.cell(r, 1, "LIVE CHECKS")
    c.font, c.fill = F_SECT, FILL_HDR
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for i, h in enumerate(["Check", "Result", "Meaning"], start=1):
        header_cell(ws, r, i, h)
    r += 1

    mrange = "%s!$A$%d:$A$%d" % (MASTER, FIRST_ROW, LAST_ROW)
    live = [
        ("Accounts in MASTER", "=COUNTA(%s)" % mrange,
         "Rows the whole workbook is built from"),
        ("Duplicate codes", "=SUMPRODUCT((%s<>\"\")*(COUNTIF(%s,%s&\"\")>1))"
         % (mrange, mrange, mrange),
         "Must be 0. A duplicate makes lookups return the wrong row."),
        ("Placeholder codes still unfixed",
         "=COUNTIF(%s,\"X_*\")" % mrange,
         "Rows that had no CODE in the old file. Replace X_* with the real code."),
        ("Accounts with no fund",
         "=SUMPRODUCT((%s<>\"\")*(%s!$G$%d:$G$%d=0))"
         % (mrange, MASTER, FIRST_ROW, LAST_ROW),
         "ROI cannot be computed for these; they show 0.00%."),
        ("Accounts with no P&L all year",
         "=SUMPRODUCT((%s<>\"\")*('%s'!$O$%d:$O$%d=0))"
         % (mrange, FY_SHEET, FIRST_ROW, LAST_ROW),
         "Dormant or never traded. Consider setting STATUS to Closed."),
    ]
    for i, (lab, f, why) in enumerate(live):
        ws.cell(r + i, 1, lab).font = F_LBL
        v = ws.cell(r + i, 2, f)
        v.number_format, v.font, v.fill, v.border = "0", F_KPI, FILL_LIGHT, BOX
        ws.cell(r + i, 3, why).font = F_BODY
        ws.cell(r + i, 1).border = BOX
    r += len(live) + 1

    for i, (qkey, sh, labels) in enumerate(QUARTERS):
        ws.cell(r, 1, "%s reconciles to its month sheets" % qkey).font = F_LBL
        ws.cell(r, 2, "=SUM('%s'!$G$%d:$L$%d)-(SUM('%s'!$G$%d:$G$%d)+"
                      "SUM('%s'!$G$%d:$G$%d)+SUM('%s'!$G$%d:$G$%d))"
                % (sh, LAST_ROW + 3, LAST_ROW + 3,
                   MONTH_SHEET[labels[0]], FIRST_ROW, LAST_ROW,
                   MONTH_SHEET[labels[1]], FIRST_ROW, LAST_ROW,
                   MONTH_SHEET[labels[2]], FIRST_ROW, LAST_ROW))
        ws.cell(r, 3, "Row %d of %s must read zero across all three months."
                % (LAST_ROW + 3, sh)).font = F_BODY
        r += 1
    r += 1

    c = ws.cell(r, 1, "FOUND DURING REBUILD  (from the old workbook)")
    c.font, c.fill = F_SECT, FILL_HDR
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for i, h in enumerate(["Severity", "Area", "Detail"], start=1):
        header_cell(ws, r, i, h)
    r += 1

    rows = []
    rows.append(["Fixed", "Broken external link",
                 "The SEP column of FY-(26-27)Q2 pointed at a missing external "
                 "SEP-2026 workbook (63 rows). Now reads the local SEP-2026 sheet."])
    rows.append(["Fixed", "Broken external link",
                 "P3361 Nishaanth and P3354 Sahil Rathore pulled August from a "
                 "missing external workbook, so Q2 showed 0 instead of 76,713 "
                 "and 2,14,254. Now read from the local AUG-2026 sheet."])
    rows.append(["Fixed", "Broken reference",
                 "The Q3 columns of the old FY-(26-27)Q1 sheet were #REF!. "
                 "Q3 now has a real sheet."])
    rows.append(["Fixed", "Fragile lookup",
                 "VLOOKUP($A$3:$A$127,...) relied on implicit intersection and "
                 "stopped at a hardcoded source row. Replaced with INDEX/MATCH "
                 "on CODE over a fixed %d-row range." % N_ROWS])
    for sheet, row, name, _id, code in placeholders:
        rows.append(["Action", "Row had no CODE",
                     "%s row %d (%s / %s) was given placeholder code %s. Its P&L "
                     "never rolled up before. Replace with the real code in MASTER."
                     % (sheet, row, name or "no name", _id or "no id", code)])
    for sev, area, detail in conflicts:
        rows.append([sev, area, detail])

    for i, (sev, area, detail) in enumerate(rows):
        ws.cell(r + i, 1, sev).font = F_BODY
        ws.cell(r + i, 2, area).font = F_BODY
        ws.cell(r + i, 3, detail).font = F_BODY
        for c_ in range(1, 4):
            ws.cell(r + i, c_).border = BOX
        if sev == "Action":
            for c_ in range(1, 4):
                ws.cell(r + i, c_).fill = PatternFill("solid", fgColor=REDBG)
        elif sev == "Check":
            for c_ in range(1, 4):
                ws.cell(r + i, c_).fill = FILL_AMBER

    for c_, w in {1: 12, 2: 24, 3: 110}.items():
        ws.column_dimensions[L(c_)].width = w
    ws.sheet_view.showGridLines = False
    return len(rows)


# =============================================================== 8. SCRATCH ==
def build_scratch(wb):
    """Preserve the ad-hoc cells that lived off to the right of the old sheets."""
    src = openpyxl.load_workbook(SRC, data_only=False)
    srcv = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.create_sheet(SCRATCH)
    title_bar(ws, "SCRATCH  -  preserved working cells",
              "Ad-hoc cells that sat outside the data grid in the old workbook, "
              "kept here with their original location so nothing was lost.", 5)
    for i, h in enumerate(["Sheet", "Cell", "Formula / text", "Value"], start=1):
        header_cell(ws, 3, i, h)

    r = 4
    for sh in SRC_MONTHS + SRC_QUARTERS:
        s, sv = src[sh], srcv[sh]
        # scratch starts after that sheet's own "<MONTH> ROI" column; the
        # columns are in a different place on every month sheet, so find it
        # rather than assuming. Quarter sheets end at N.
        limit = 14
        if sh in SRC_MONTHS:
            lab = sh.split("-")[0]
            for c in range(1, s.max_column + 1):
                if norm(s.cell(2, c).value).upper() == lab + " ROI":
                    limit = c
                    break
        for row in s.iter_rows():
            for cell in row:
                if cell.value in (None, ""):
                    continue
                if cell.column <= limit:
                    continue
                # Q1's old P..Z block was the Q2/Q3/Q4/FY rollup, now rebuilt
                # as real sheets. Its Q3 columns were #REF! anyway.
                if sh in SRC_QUARTERS and cell.column <= 26:
                    continue
                ws.cell(r, 1, sh).font = F_BODY
                ws.cell(r, 2, "%s%d" % (L(cell.column), cell.row)).font = F_BODY
                ws.cell(r, 3, str(cell.value)).font = F_BODY
                v = sv.cell(cell.row, cell.column).value
                ws.cell(r, 4, v if isinstance(v, (int, float, str)) else None).font = F_BODY
                for c_ in range(1, 5):
                    ws.cell(r, c_).border = BOX
                r += 1
    for c_, w in {1: 16, 2: 8, 3: 46, 4: 16}.items():
        ws.column_dimensions[L(c_)].width = w
    ws.sheet_view.showGridLines = False
    return r - 4


# =================================================================== DRIVER ==
def main():
    print("Reading %s ..." % SRC)
    accounts, order, daily, fund_by_month, conflicts, placeholders = read_source()
    print("  %d accounts, %d daily P&L cells, %d rows had no CODE"
          % (len(order), len(daily), len(placeholders)))
    if len(order) > N_ROWS:
        sys.exit("More accounts (%d) than provisioned rows (%d) -- raise N_ROWS."
                 % (len(order), N_ROWS))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary(wb, accounts)
    build_master(wb, accounts, order, fund_by_month)
    carried = 0
    for lab, sheet, year, month in MONTHS:
        carried += build_month(wb, lab, sheet, year, month, daily, order, accounts)
    for qkey, sheet, labels in QUARTERS:
        build_quarter(wb, qkey, sheet, labels)
    build_fy(wb)
    n_issues = build_issues(wb, conflicts, placeholders, accounts)
    n_scratch = build_scratch(wb)

    wb.active = 0
    wb.save(OUT)
    print("  carried %d daily values into the new month sheets" % carried)
    print("  %d issue rows, %d scratch cells preserved" % (n_issues, n_scratch))
    print("Wrote %s  (%d sheets)" % (OUT, len(wb.sheetnames)))
    print("  " + " | ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
