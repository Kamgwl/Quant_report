import openpyxl
import re

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["FY-(26-27)Q1"]

COL_CODE     = 2
COL_NAME     = 3
COL_STRATEGY = 5
COL_SEGMENT  = 6
COL_FUND     = 7
COL_APR      = 8
COL_MAY      = 10
COL_JUN      = 12

def safe_int(v, default=0):
    try:
        return int(round(float(v))) if v is not None else default
    except (ValueError, TypeError):
        return default

def is_data_row(ws, r):
    code = ws.cell(r, COL_CODE).value
    if code is None:
        return False
    code = str(code).strip()
    if code.upper() in ("CODE", "", "NONE"):
        return False
    if not re.match(r'^[A-Z0-9_]{3,10}$', code.upper()):
        return False
    return True

fo_apr = 0; fo_may = 0; fo_jun = 0
cash_apr = 0; cash_may = 0; cash_jun = 0

print("Accounts classified as CASH/ETF:")
print("-" * 60)
for r in range(4, ws.max_row + 1):
    if not is_data_row(ws, r):
        continue
    
    code = str(ws.cell(r, COL_CODE).value).strip()
    name = str(ws.cell(r, COL_NAME).value or "").strip()
    segment = str(ws.cell(r, COL_SEGMENT).value or "").strip().upper()
    
    apr = safe_int(ws.cell(r, COL_APR).value)
    may = safe_int(ws.cell(r, COL_MAY).value)
    jun = safe_int(ws.cell(r, COL_JUN).value)
    
    is_cash = segment in ("CASH", "ETF")
    if is_cash:
        print(f"Row {r:>2}: {code} - {name} (Segment: {segment}) | Apr={apr:>9} | May={may:>9} | Jun={jun:>9}")
        cash_apr += apr
        cash_may += may
        cash_jun += jun
    else:
        fo_apr += apr
        fo_may += may
        fo_jun += jun

print("\nSums based on Segment column:")
print(f"F&O April P&L: {fo_apr:,}")
print(f"F&O May P&L  : {fo_may:,}")
print(f"F&O June P&L : {fo_jun:,}")
print(f"Cash April P&L: {cash_apr:,}")
print(f"Cash May P&L  : {cash_may:,}")
print(f"Cash June P&L : {cash_jun:,}")
