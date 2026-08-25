import openpyxl

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["APR-2026"]

print("APR-2026 Column Headers (Row 3):")
for col in range(1, ws.max_column + 1):
    val = ws.cell(3, col).value
    if val:
        print(f"Col {col}: {val}")

# Also check last columns of a data row (e.g. row 4)
print("\nRow 4 end values:")
for col in range(ws.max_column - 10, ws.max_column + 1):
    val = ws.cell(4, col).value
    header = ws.cell(3, col).value
    print(f"Col {col} ({header}): {val}")
