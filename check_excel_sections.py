import openpyxl
import re

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["FY-(26-27)Q1"]

COL_CODE     = 2
COL_NAME     = 3
COL_STRATEGY = 5
COL_SEGMENT  = 6
COL_FUND     = 7
COL_APR      = 8
COL_MAY      = 10
COL_JUN      = 12

def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default

def safe_int(v, default=0):
    try:
        return int(round(float(v))) if v is not None else default
    except (ValueError, TypeError):
        return default

def is_data_row(ws, r):
    code = ws.cell(r, COL_CODE).value
    if code is None:
        return False
    code = str(code).strip()
    if code.upper() in ("CODE", "", "NONE"):
        return False
    if not re.match(r'^[A-Z0-9_]{3,10}$', code.upper()):
        return False
    return True

# Group 1: Rows 4 to 66 (Normal Section)
g1_fund = 0; g1_apr = 0; g1_may = 0; g1_jun = 0
for r in range(4, 67):
    if not is_data_row(ws, r):
        continue
    g1_fund += safe_float(ws.cell(r, COL_FUND).value)
    g1_apr += safe_int(ws.cell(r, COL_APR).value)
    g1_may += safe_int(ws.cell(r, COL_MAY).value)
    g1_jun += safe_int(ws.cell(r, COL_JUN).value)

# Group 2: Rows 70 to 72 (Deepesh Group Section)
g2_fund = 0; g2_apr = 0; g2_may = 0; g2_jun = 0
for r in range(70, 73):
    if not is_data_row(ws, r):
        continue
    g2_fund += safe_float(ws.cell(r, COL_FUND).value)
    g2_apr += safe_int(ws.cell(r, COL_APR).value)
    g2_may += safe_int(ws.cell(r, COL_MAY).value)
    g2_jun += safe_int(ws.cell(r, COL_JUN).value)

print("--- EXCEL SECTIONS ---")
print(f"Normal Section (Rows 4-66) Totals:")
print(f"  Fund: {g1_fund} Cr")
print(f"  Apr P&L: {g1_apr:,}")
print(f"  May P&L: {g1_may:,}")
print(f"  Jun P&L: {g1_jun:,}")
print(f"Deepesh Group Section (Rows 70-72) Totals:")
print(f"  Fund: {g2_fund} Cr")
print(f"  Apr P&L: {g2_apr:,}")
print(f"  May P&L: {g2_may:,}")
print(f"  Jun P&L: {g2_jun:,}")
