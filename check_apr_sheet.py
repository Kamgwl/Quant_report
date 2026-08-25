import openpyxl

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["APR-2026"]

print(f"APR-2026 sheet max row: {ws.max_row}, max col: {ws.max_column}")
for r in range(1, min(ws.max_row + 1, 100)):
    row_vals = [ws.cell(r, c).value for c in range(1, 15)]
    # Print rows that have at least one non-None value
    if any(val is not None for val in row_vals):
        print(f"Row {r:>2}: {row_vals}")
