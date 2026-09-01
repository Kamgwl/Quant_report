"""
update_dashboard.py
───────────────────
Reads "Quant Strategy(2026-27).xlsx"  (sheet: FY-(26-27)Q1)
and updates the RAW data array in:
  • index.html                       (standalone CDN page → GitHub Pages)
  • indx.html                        (JSX source)
  • quant-dashboard/src/data.js      (Vite project module)

Run from  d:\\file_dash  then deploy.bat does the git push.
"""

import openpyxl
import re
import sys
import os
from datetime import datetime

# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_FILE  = "Quant Strategy(2026-27).xlsx"
SHEET_Q1    = "FY-(26-27)Q1"   # Apr / May / Jun
SHEET_Q2    = "FY-(26-27)Q2"   # Jul / Aug / Sep

# Column positions (1-based). The Q1 and Q2 sheets share an identical layout;
# only the three month columns mean different months (Apr-Jun vs Jul-Sep).
# NOTE: the 2026-27 workbook dropped the old leading blank column A, so every
# field shifted one column left vs. the original layout (code was col 2 → now
# col 1). Column 3 is now an "ID" column (e.g. ATS / XTS09) which we ignore.
COL_CODE     = 1
COL_NAME     = 2
COL_STRATEGY = 4
COL_SEGMENT  = 5
COL_FUND     = 6
COL_M1       = 7    # Apr (Q1)  /  Jul (Q2)
COL_M1_ROI   = 8
COL_M2       = 9    # May (Q1)  /  Aug (Q2)
COL_M2_ROI   = 10
COL_M3       = 11   # Jun (Q1)  /  Sep (Q2)
COL_M3_ROI   = 12

DATA_START_ROW = 4   # row 3 is the header row; data from row 4

# Each quarter sheet's month columns are VLOOKUPs into that month's detail sheet
# (APR-2026, ... AUG-2026), matched on the label in row 2 ("AUG" / "AUG ROI").
# The detail sheet carries the same label over its own monthly total column.
#
# Some rollup rows point at an EXTERNAL workbook copy of the detail sheet
# ('[2]AUG-2026'). Excel's IFERROR swallows the broken link and returns 0, so
# those accounts silently vanish from the quarter even though the local detail
# sheet has their numbers. Fall back to the local sheet whenever that happens.
MONTH_HEADER_ROW = 2

# NAME_OVERRIDES:  "CODE": (display_name, strategy, fund_override_or_None)
# fund_override is in Crores; set None to use value from Excel.
NAME_OVERRIDES = {
    # code          display_name              strategy         fund_override
    "P3224":  ("Ansh Cash",           "CASH",          None),
    "P2827":  ("Bharat ETF",          "ETF",           None),
    "P2827_2":("Bharat Cash",         "CASH",          None),
    "P3090":  ("Bharat Straddle",     "STRADDLE",      None),
    "P2954":  ("Bharat Munjal_5",     "STRADDLE",      None),
    "P3196":  ("Bharat Munjal BFO",   "STRADDLE",      None),
    "P2954_2":("Bharat Munjal S1515", "STRADDLE",      None),
    "P3079":  ("Yogesh Cash",         "CASH",          None),
    "P2777":  ("Yogesh Kumar",        "STRADDLE",      None),
    "P2826":  ("Yogesh ATS",          "",              None),
    "P3039":  ("Abhishek",            "",              None),
    "P2967":  ("Aman",                "Money Circle",  None),
    "P2951":  ("Aman 20-60",          "20_60",         None),
    "P3166":  ("Aman Dohre",          "NSE FO",        None),
    "P3390":  ("Aman Dohre XTS34",    "NSE FO",        None),
    "P3323":  ("Amit&Kartik",         "NSE FO",        None),
    "P3323_1":("Amit&Kartik XTS35",   "",              None),
    "PSW023": ("Ansh SWV",            "Falcon",        None),
    "PHPO06": ("Ansh HPO",            "Falcon",        None),
    "PSW019": ("Kartaram SWV097",     "Spider 2.0",    None),
    "SWV0096":("Kartaram SWV096",     "Spider 2.0",    None),
    "PHPO07": ("Kartaram HPO",        "Spider 2.0",    None),
    "XMR0548":("Kartaram XMR",        "Spider 2.0",    None),
    "PA528":  ("Archana Ma'am",       "SPIDER 2.0",    None),
    "P3181":  ("Harshit XTS32",       "SELF",          None),
    "P3240":  ("Harshit XTS38",       "SELF",          None),
    "P3311":  ("Himanshu Pal",        "SELF",          None),
    "P3109":  ("Jyoti Prakesh XTS30", "Straddle",      None),
    "P3313":  ("Jyoti Prakesh ATS",   "TRANDING",      None),
    "P3146":  ("Jinesh Jain",         "STOCK FO",      None),
    "P2105":  ("Kamlesh",             "Wanda",         None),
    "P3024":  ("Kartaram_1",          "Spider 2.0",    None),
    "P3202":  ("Gagandeep",           "SELF",          None),
    "P3020":  ("Mahavir Jindal",      "SELF",          None),
    "P3218":  ("Muhunthan",           "CASH",          None),
    "P3341":  ("Maneesh Yadav",       "NSE FO",        None),
    "P2971":  ("Neeraj Garg_1",       "Straddle",      None),
    "P2971_2":("Neeraj Garg_2",       "ACE1",          None),
    "P2940":  ("Prince Cash",         "cash",          None),
    "P3347":  ("Prince FO",           "FO",            None),
    "P3070":  ("Prateek_2",           "multipul",      None),
    "P2999":  ("Prateek ETF",         "ETF",           None),
    "P3186":  ("Prateek NFO",         "NFO",           None),
    # P3135 fund is 3 Cr in reality; Excel shows 1 Cr (data entry error)
    "P3135":  ("Piyush Singhal",      "Strangle",      3),
    "PSW024": ("Piyush SWV",          "Strangle",      None),
    "P3168":  ("Prabjot Singh",       "NSE FO",        None),
    "P3342":  ("Raghav Tuli",         "",              None),
    "P3082":  ("Ramakar Jha",         "STOCK FO",      None),
    "P3048":  ("Shahid",              "SELF",          None),
    "P3208":  ("Sajal Sharma ATS",    "",              None),
    "P3335":  ("Sajal Sharma_2",      "",              None),
    "P3119":  ("Sajal Sharma_3",      "",              None),
    "P3297":  ("Sudeep",              "ATS",           None),
    "P3360":  ("Vidhya Sagar",        "SELF",          None),
    "P3385":  ("Vikas Gupta XTS13",   "SELF",          None),
    "P3386":  ("Vikas Gupta XTS15",   "SELF",          None),
    "P3334":  ("Varun Tondan XTS10",  "self",          None),
    "P3110":  ("Varun Tondan",        "self",          None),
    "P3113":  ("Vaishali",            "self",          None),
    "P2817":  ("DK Sir_M",            "Maximum",       None),
    "P2792":  ("DK Sir_T",            "ThanosNF",      None),
    "P3013":  ("Deepesh_1",           "NA",            None),
    "P3112":  ("Deepesh_2",           "NA",            None),
    "P3361":  ("Nishaanth",           "NA",            0),    # no fund denominator
}

DEEPESH_CODES = {"P3013", "P3112", "P3361"}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default

def safe_int(v, default=0):
    try:
        return int(round(float(v))) if v is not None else default
    except (ValueError, TypeError):
        return default

def fmt_num(v):
    if v == 0:
        return "0"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(round(v, 4))

def is_data_row(ws, r):
    code = ws.cell(r, COL_CODE).value
    if code is None:
        return False
    code = str(code).strip()
    if code.upper() in ("CODE", "", "NONE"):
        return False
    if not re.match(r'^[A-Z0-9_]{3,10}$', code.upper()):
        return False
    return True

# ── EXTRACTION ────────────────────────────────────────────────────────────────
def read_quarter(ws):
    """Read one quarterly sheet into {code: {name, strategy, segment, fund, m1..m3(_roi)}}.
    m1/m2/m3 are the three months in sheet order (Apr-Jun for Q1, Jul-Sep for Q2)."""
    out = {}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if not is_data_row(ws, r):
            continue
        code = str(ws.cell(r, COL_CODE).value).strip()
        out[code] = {
            "name":     str(ws.cell(r, COL_NAME).value or "").strip(),
            "strategy": str(ws.cell(r, COL_STRATEGY).value or "").strip(),
            "segment":  str(ws.cell(r, COL_SEGMENT).value or "").strip(),
            "fund":     safe_float(ws.cell(r, COL_FUND).value),
            "m1":       safe_int(ws.cell(r, COL_M1).value),
            "m1_roi":   safe_float(ws.cell(r, COL_M1_ROI).value),
            "m2":       safe_int(ws.cell(r, COL_M2).value),
            "m2_roi":   safe_float(ws.cell(r, COL_M2_ROI).value),
            "m3":       safe_int(ws.cell(r, COL_M3).value),
            "m3_roi":   safe_float(ws.cell(r, COL_M3_ROI).value),
        }
    return out

def month_total_cols(ws, want):
    """(pnl_col, roi_col) for a month detail sheet.

    Prefer an exact row-2 label match ("SEP" / "SEP ROI"). Failing that, fall
    back to the sheet's structure: every month sheet carries exactly one
    "<X>" / "<X> ROI" pair to the right of the dated columns, whatever <X>
    happens to say. SEP-2026 was copied from AUG-2026 and still labels its
    total columns "AUG" / "AUG ROI", so trusting the label alone loses the
    whole month — and worse, hands September's numbers to whoever asks for
    August."""
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(MONTH_HEADER_ROW, c).value
        if isinstance(v, str) and v.strip():
            hdr[v.strip().upper()] = c
    if want in hdr:
        return hdr[want], hdr.get(want + " ROI")
    for text, c in hdr.items():
        if text.endswith(" ROI") and text[:-4].strip() in hdr:
            return hdr[text[:-4].strip()], c
    return None, None


def read_month_detail(wb, label):
    """(sheet_name, {code: (pnl, roi, n_rows)}) for one month.

    The sheet is located by NAME ("SEP" -> "SEP-2026"), not by trusting the
    labels inside it, because a sheet copied from the previous month keeps the
    previous month's labels. Only if no sheet name matches do we fall back to
    scanning for the label.

    An account can occupy more than one row (AUG-2026 and JULY-2026 both hold
    two rows coded P3390). Values are SUMMED across them, and n_rows reports how
    many were folded together — the rollup's VLOOKUP only ever sees the first."""
    want = label.strip().upper()

    target = None
    for ws in wb.worksheets:
        if ws.title.startswith("FY-"):
            continue
        if ws.title.strip().upper().startswith(want + "-"):
            target = ws
            break
    if target is None:
        for ws in wb.worksheets:
            if ws.title.startswith("FY-"):
                continue
            c_pnl, _ = month_total_cols(ws, want)
            if c_pnl is not None and want in {
                    str(ws.cell(MONTH_HEADER_ROW, c).value).strip().upper()
                    for c in range(1, ws.max_column + 1)}:
                target = ws
                break
    if target is None:
        return None, {}

    c_pnl, c_roi = month_total_cols(target, want)
    if c_pnl is None:
        return target.title, {}

    out = {}
    for r in range(DATA_START_ROW, target.max_row + 1):
        if not is_data_row(target, r):
            continue
        code = str(target.cell(r, COL_CODE).value).strip()
        pnl = safe_int(target.cell(r, c_pnl).value)
        roi = safe_float(target.cell(r, c_roi).value) if c_roi else 0.0
        prev = out.get(code, (0, 0.0, 0))
        out[code] = (prev[0] + pnl, prev[1] + roi, prev[2] + 1)
    return target.title, out


def repair_from_detail(wb, ws_q, rows):
    """Reconcile the quarter rollup against the local monthly detail sheets.

    Two defects are corrected, both of them cases where the rollup's VLOOKUP
    does not reflect what the detail sheet actually holds:

      LINK  the rollup formula points at an external workbook copy of the sheet
            ('[2]AUG-2026'), IFERROR swallows the dead link and yields 0
      MERGE the account occupies several rows and VLOOKUP returns only the first

    Nothing else is touched, so a rollup value that legitimately differs is left
    alone and the whole pass becomes a no-op once the workbook is repaired."""
    fixed = 0
    for slot, col in (("m1", COL_M1), ("m2", COL_M2), ("m3", COL_M3)):
        label = ws_q.cell(MONTH_HEADER_ROW, col).value
        if not isinstance(label, str):
            continue
        sheet, detail = read_month_detail(wb, label)
        if not detail:
            continue
        for code, rec in rows.items():
            pnl, roi, n_rows = detail.get(code, (0, 0.0, 0))
            broken_link = rec[slot] == 0 and pnl != 0
            merged_rows = n_rows > 1
            if not (broken_link or merged_rows):
                continue

            if merged_rows:
                # Summed ROIs are meaningless across rows carrying different
                # fund bases — recompute against the account's own fund.
                fund = rows[code].get("fund") or 0
                roi = (pnl / (fund * 1e7) * 100) if fund else roi
                kind, note = "MERGE", f"{n_rows} rows -> 1"
            else:
                kind, note = "LINK", "rollup 0"

            was = rec[slot]
            rec[slot], rec[slot + "_roi"] = pnl, roi
            fixed += 1
            print(f"    [{kind}] {code:<9} {label.strip():<5} {note:<12} "
                  f"{was:>12,} -> {pnl:>12,}  (from {sheet})")
    return fixed


def extract_accounts(excel_path):
    print(f"[1/4] Reading {os.path.basename(excel_path)} ...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    q1 = read_quarter(wb[SHEET_Q1])
    q2 = read_quarter(wb[SHEET_Q2]) if SHEET_Q2 in wb.sheetnames else {}
    print(f"    Q1 sheet: {len(q1)} accounts | Q2 sheet: {len(q2)} accounts.")

    n = repair_from_detail(wb, wb[SHEET_Q1], q1)
    if SHEET_Q2 in wb.sheetnames:
        n += repair_from_detail(wb, wb[SHEET_Q2], q2)
    print(f"    Reconciled {n} month value(s) against the detail sheets.")

    # Union of codes, Q1 order first, then any Q2-only codes appended.
    codes = list(q1.keys()) + [c for c in q2 if c not in q1]
    EMPTY = {"m1": 0, "m1_roi": 0.0, "m2": 0, "m2_roi": 0.0, "m3": 0, "m3_roi": 0.0}

    accounts = []
    for code in codes:
        a1  = q1.get(code, EMPTY)
        a2  = q2.get(code, EMPTY)
        ident = q1.get(code) or q2.get(code)   # identity/fund from Q1 if present

        raw_name  = ident["name"]
        raw_strat = ident["strategy"]
        raw_seg   = ident["segment"]
        fund      = ident["fund"]

        if code in NAME_OVERRIDES:
            ov = NAME_OVERRIDES[code]
            disp_name = ov[0]
            strat     = ov[1]
            if ov[2] is not None:
                fund = ov[2]
        else:
            disp_name = raw_name
            strat     = raw_strat

        is_deepesh = code in DEEPESH_CODES
        group = 2 if is_deepesh else 1

        accounts.append({
            "code":       code,
            "name":       disp_name,
            "strategy":   strat,
            "segment":    raw_seg,
            "group":      group,
            "fund":       fund,
            # Q1 months
            "apr": a1["m1"], "apr_roi": round(a1["m1_roi"], 4),
            "may": a1["m2"], "may_roi": round(a1["m2_roi"], 4),
            "jun": a1["m3"], "jun_roi": round(a1["m3_roi"], 4),
            # Q2 months
            "jul": a2["m1"], "jul_roi": round(a2["m1_roi"], 4),
            "aug": a2["m2"], "aug_roi": round(a2["m2_roi"], 4),
            "sep": a2["m3"], "sep_roi": round(a2["m3_roi"], 4),
            "is_deepesh": is_deepesh,
        })

    print(f"    Merged {len(accounts)} accounts (Q1 + Q2).")
    return accounts

# ── JS RAW ARRAY BUILDER ──────────────────────────────────────────────────────
def build_raw_js(accounts):
    lines = []
    in_deepesh = False
    for i, acc in enumerate(accounts):
        comma = "," if i < len(accounts) - 1 else ""

        if acc["is_deepesh"] and not in_deepesh:
            lines.append("// Deepesh group")
            in_deepesh = True

        # Escape single quotes in name (JS string uses double quotes so safe)
        name_js  = acc["name"].replace('"', '\\"')
        strat_js = acc["strategy"].replace('"', '\\"')
        seg_js   = acc["segment"].replace('"', '\\"')

        line = (
            f'{{ code:"{acc["code"]}", name:"{name_js}", strategy:"{strat_js}", segment:"{seg_js}", group:{acc["group"]}, '
            f'fund:{fmt_num(acc["fund"])}, '
            f'apr:{acc["apr"]}, apr_roi:{fmt_num(acc["apr_roi"])}, '
            f'may:{acc["may"]}, may_roi:{fmt_num(acc["may_roi"])}, '
            f'jun:{acc["jun"]}, jun_roi:{fmt_num(acc["jun_roi"])}, '
            f'jul:{acc["jul"]}, jul_roi:{fmt_num(acc["jul_roi"])}, '
            f'aug:{acc["aug"]}, aug_roi:{fmt_num(acc["aug_roi"])}, '
            f'sep:{acc["sep"]}, sep_roi:{fmt_num(acc["sep_roi"])} }}{comma}'
        )
        lines.append(line)

    return lines

# ── FILE UPDATER ──────────────────────────────────────────────────────────────
# Matches both:   const RAW = [...]
#            and: export const RAW = [...]
RAW_PATTERN = re.compile(
    r'((?:export\s+)?const\s+RAW\s*=\s*\[)(.*?)(\];)',
    re.DOTALL
)

def update_file(filepath, new_raw_lines, label, is_data_js=False):
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    raw_body = "\n".join(new_raw_lines)

    if is_data_js and not RAW_PATTERN.search(content):
        # Rebuild data.js from scratch
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_content = (
            f'// Raw account data from "{EXCEL_FILE}" (Q1: Apr-Jun 2026).\n'
            f'// Auto-updated: {ts}\n'
            f'export const RAW = [\n{raw_body}\n];\n'
        )
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(new_content)
        print(f"    Rebuilt {label} ({len(new_raw_lines)} data lines).")
        return True

    new_content, count = RAW_PATTERN.subn(
        lambda m: f"{m.group(1)}\n{raw_body}\n{m.group(3)}",
        content,
        count=1
    )
    if count == 0:
        print(f"    [WARN] RAW array not found in {label} — skipping.")
        return False

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(new_content)
    print(f"    Updated {label} ({len(new_raw_lines)} data lines).")
    return True

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    base = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base, EXCEL_FILE)

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found:\n  {excel_path}")
        sys.exit(1)

    # 1. Extract
    accounts = extract_accounts(excel_path)

    # 2. Build JS
    print("[2/4] Building JS RAW array ...")
    raw_lines = build_raw_js(accounts)

    # 3. Update files
    print("[3/4] Updating dashboard files ...")
    targets = [
        ("index.html",                                    False),
        ("indx.html",                                     False),
        (os.path.join("quant-dashboard", "src", "data.js"), True),
    ]
    updated = 0
    for relpath, is_data_js in targets:
        full = os.path.join(base, relpath)
        label = os.path.basename(relpath)
        if not os.path.exists(full):
            print(f"    [SKIP] {label} not found.")
            continue
        if update_file(full, raw_lines, label, is_data_js):
            updated += 1

    # 4. Summary
    tot = lambda k: sum(a[k] for a in accounts)
    total_apr, total_may, total_jun = tot("apr"), tot("may"), tot("jun")
    total_jul, total_aug, total_sep = tot("jul"), tot("aug"), tot("sep")
    total_q1 = total_apr + total_may + total_jun
    total_q2 = total_jul + total_aug + total_sep
    total_fy = total_q1 + total_q2

    print(f"\n[4/4] Done -- {updated} file(s) updated.  {len(accounts)} accounts.")
    print(f"      Apr P&L: Rs. {total_apr:>14,.0f}")
    print(f"      May P&L: Rs. {total_may:>14,.0f}")
    print(f"      Jun P&L: Rs. {total_jun:>14,.0f}")
    print(f"      Q1  P&L: Rs. {total_q1:>14,.0f}")
    print(f"      ---")
    print(f"      Jul P&L: Rs. {total_jul:>14,.0f}")
    print(f"      Aug P&L: Rs. {total_aug:>14,.0f}")
    print(f"      Sep P&L: Rs. {total_sep:>14,.0f}")
    print(f"      Q2  P&L: Rs. {total_q2:>14,.0f}")
    print(f"      ===")
    print(f"      FY  P&L: Rs. {total_fy:>14,.0f}")

if __name__ == "__main__":
    main()
