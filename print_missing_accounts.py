import openpyxl

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["FY-(26-27)Q1"]

missing_codes = ["P3323_1", "P2827_2", "P2954_2", "P2971_2"]

print("Missing accounts details in Excel:")
for r in range(4, ws.max_row + 1):
    code = ws.cell(r, 2).value
    if code in missing_codes:
        row_vals = [ws.cell(r, c).value for c in range(1, 15)]
        print(f"Row {r}: {row_vals}")
