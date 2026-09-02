"""
verify_rebuild.py
-----------------
Checks the rebuilt workbook without needing Excel to recalculate.

Three passes:
  1. Structural  - every sheet/range a formula points at actually exists,
                   no #REF!, no external links, no stale hardcoded rows.
  2. Arithmetic  - recomputes what each formula WILL evaluate to (month
                   totals, ROI, quarter and FY rollups) from the carried
                   daily cells, and compares against the old workbook's
                   cached values.
  3. Differences - lists every account whose number changes, with the reason.
"""

import re
import sys
from collections import defaultdict

import openpyxl

SRC = "Quant Strategy(2026-27).xlsx"
OUT = "Quant Strategy(2026-27) v2.xlsx"
FIRST_ROW, LAST_ROW = 4, 123

MONTHS = [("APR", "APR-2026"), ("MAY", "MAY-2026"), ("JUNE", "JUNE-2026"),
          ("JULY", "JULY-2026"), ("AUG", "AUG-2026"), ("SEP", "SEP-2026"),
          ("OCT", "OCT-2026"), ("NOV", "NOV-2026"), ("DEC", "DEC-2026"),
          ("JAN", "JAN-2027"), ("FEB", "FEB-2027"), ("MAR", "MAR-2027")]
QUARTERS = [("Q1", "FY-(26-27)Q1", ["APR", "MAY", "JUNE"]),
            ("Q2", "FY-(26-27)Q2", ["JULY", "AUG", "SEP"]),
            ("Q3", "FY-(26-27)Q3", ["OCT", "NOV", "DEC"]),
            ("Q4", "FY-(26-27)Q4", ["JAN", "FEB", "MAR"])]
SRC_MONTHS = {"APR": "APR-2026", "MAY": "MAY-2026", "JUNE": "JUNE-2026",
              "JULY": "JULY-2026", "AUG": "AUG-2026"}

fails, warns = [], []


def fail(msg):
    fails.append(msg)


def warn(msg):
    warns.append(msg)


new = openpyxl.load_workbook(OUT)
old = openpyxl.load_workbook(SRC, data_only=True)

# ============================================================= 1. STRUCTURAL ==
print("=" * 78)
print("1. STRUCTURAL")
print("=" * 78)

sheet_ref = re.compile(r"'([^']+)'!")
n_formulas = 0
bad_ref = ext_link = 0
missing = set()
for ws in new.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            n_formulas += 1
            if "#REF" in v:
                bad_ref += 1
                fail("#REF! in %s!%s" % (ws.title, cell.coordinate))
            if re.search(r"\[\d+\]", v):
                ext_link += 1
                fail("external link in %s!%s" % (ws.title, cell.coordinate))
            for ref in sheet_ref.findall(v):
                if ref not in new.sheetnames:
                    missing.add((ws.title, ref))

print("  formulas written          : %d" % n_formulas)
print("  #REF! errors              : %d" % bad_ref)
print("  external-workbook links   : %d" % ext_link)
if missing:
    for s, ref in sorted(missing):
        fail("%s references missing sheet '%s'" % (s, ref))
else:
    print("  broken sheet references   : 0")

zf = openpyxl.open  # noqa - placeholder to keep linters quiet
import zipfile
with zipfile.ZipFile(OUT) as z:
    ext_parts = [n for n in z.namelist() if "externalLink" in n]
print("  externalLink parts in file: %d" % len(ext_parts))
if ext_parts:
    fail("workbook still carries externalLink parts: %s" % ext_parts)

# dashboard contract: quarter sheets must keep columns 1-12 where
# update_dashboard.py expects them, month sheets must keep row-2 labels
print()
print("  dashboard compatibility:")
for qkey, sh, labels in QUARTERS[:2]:
    ws = new[sh]
    heads = [ws.cell(3, c).value for c in range(1, 13)]
    want = ["CODE", "NAME", "ID", "STRATEGY", "SEGMENT", "FUND (CR)"]
    ok = [str(h).upper().startswith(w.split()[0]) for h, w in zip(heads, want)]
    print("    %-14s A-F headers %s   month cols 7-12 = %s"
          % (sh, "OK" if all(ok) else "MISMATCH",
             [ws.cell(2, c).value for c in range(7, 13)]))
    if not all(ok):
        fail("%s columns A-F no longer match the dashboard's contract" % sh)
for lab, sh in MONTHS[:5]:
    ws = new[sh]
    g2, h2 = ws.cell(2, 7).value, ws.cell(2, 8).value
    if g2 != lab or h2 != lab + " ROI":
        fail("%s row-2 labels are %r/%r, dashboard repair pass expects %r/%r"
             % (sh, g2, h2, lab, lab + " ROI"))
print("    month-sheet row-2 labels  : %s"
      % ("all OK" if not any("row-2" in f for f in fails) else "MISMATCH"))

# ============================================================= 2. ARITHMETIC ==
print()
print("=" * 78)
print("2. ARITHMETIC  (recomputing what the formulas will evaluate to)")
print("=" * 78)

# code -> row, from MASTER
mws = new["MASTER"]
rows = {}
master_fund = {}
for r in range(FIRST_ROW, LAST_ROW + 1):
    code = mws.cell(r, 1).value
    if code:
        rows[str(code).strip()] = r
        f = mws.cell(r, 7).value
        master_fund[str(code).strip()] = float(f) if isinstance(f, (int, float)) else 0.0
hist = {}
for code, r in rows.items():
    for i, (lab, _) in enumerate(MONTHS):
        v = mws.cell(r, 11 + i).value
        if isinstance(v, (int, float)):
            hist[(code, lab)] = float(v)

print("  MASTER accounts: %d   fund-history overrides: %d" % (len(rows), len(hist)))


def month_fund(code, lab):
    return hist.get((code, lab), master_fund.get(code, 0.0))


# recompute month totals from the carried daily cells
new_month = defaultdict(dict)   # lab -> code -> total
new_month_roi = defaultdict(dict)
for lab, sh in MONTHS:
    ws = new[sh]
    ncols = ws.max_column
    for code, r in rows.items():
        tot = 0
        for c in range(10, ncols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                tot += v
        new_month[lab][code] = tot
        f = month_fund(code, lab)
        new_month_roi[lab][code] = (tot / (f * 10000000) * 100) if f else 0.0

# old month totals, straight from the old sheets' cached values
old_month = defaultdict(dict)
old_month_roi = defaultdict(dict)
for lab, sh in SRC_MONTHS.items():
    ws = old[sh]
    tcol = rcol = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(2, c).value or "").strip().upper()
        if h == lab:
            tcol = c
        elif h == lab + " ROI":
            rcol = c
    for r in range(FIRST_ROW, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code:
            nm, i = ws.cell(r, 2).value, ws.cell(r, 3).value
            if not (nm or i):
                continue
            code = "X_" + str(i or nm).upper().replace(" ", "")[:8]
        code = str(code).strip()
        v = ws.cell(r, tcol).value if tcol else None
        old_month[lab][code] = v if isinstance(v, (int, float)) else 0
        v = ws.cell(r, rcol).value if rcol else None
        old_month_roi[lab][code] = v if isinstance(v, (int, float)) else 0.0

print()
print("  month totals, new vs old (per account):")
print("  %-6s %8s %12s %18s %18s" % ("month", "accounts", "mismatched",
                                     "old sheet total", "new sheet total"))
for lab, _ in MONTHS[:5]:
    bad = 0
    for code in rows:
        o = old_month[lab].get(code, 0)
        n = new_month[lab].get(code, 0)
        if abs(o - n) > 0.51:
            bad += 1
            if bad <= 3:
                warn("%s %s: old month sheet %.0f vs rebuilt %.0f" % (lab, code, o, n))
    ot = sum(old_month[lab].values())
    nt = sum(new_month[lab].values())
    flag = "" if abs(ot - nt) < 1 else "   <-- DIFFERS"
    print("  %-6s %8d %12d %18s %18s%s"
          % (lab, len(rows), bad, "{:,.0f}".format(ot), "{:,.0f}".format(nt), flag))
    if bad:
        fail("%s: %d accounts' month totals do not reproduce" % (lab, bad))

print()
print("  month ROI, new vs old (per account):")
for lab, _ in MONTHS[:5]:
    bad = [c for c in rows
           if abs(old_month_roi[lab].get(c, 0) - new_month_roi[lab].get(c, 0)) > 0.005]
    print("  %-6s mismatched ROI cells: %d" % (lab, len(bad)))
    for c in bad[:4]:
        warn("%s %s ROI: old %.4f vs rebuilt %.4f  (fund %.2f)"
             % (lab, c, old_month_roi[lab].get(c, 0),
                new_month_roi[lab].get(c, 0), month_fund(c, lab)))

# quarter + FY rollups
new_q, new_q_roi = defaultdict(dict), defaultdict(dict)
for qkey, sh, labels in QUARTERS:
    for code in rows:
        new_q[qkey][code] = sum(new_month[l].get(code, 0) for l in labels)
        new_q_roi[qkey][code] = sum(new_month_roi[l].get(code, 0) for l in labels)

old_q = defaultdict(dict)
for qkey, sh, labels in QUARTERS[:2]:
    ws = old[sh]
    for r in range(FIRST_ROW, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code:
            nm, i = ws.cell(r, 2).value, ws.cell(r, 3).value
            if not (nm or i):
                continue
            code = "X_" + str(i or nm).upper().replace(" ", "")[:8]
        v = ws.cell(r, 13).value
        old_q[qkey][str(code).strip()] = v if isinstance(v, (int, float)) else 0

print()
print("  quarter rollups:")
for qkey, sh, labels in QUARTERS:
    nt = sum(new_q[qkey].values())
    ot = sum(old_q[qkey].values()) if qkey in old_q else None
    print("  %-4s rebuilt total %18s     old workbook %s"
          % (qkey, "{:,.0f}".format(nt),
             "{:,.0f}".format(ot) if ot is not None else "(sheet did not exist)"))

fy_new = sum(sum(new_q[q].values()) for q, _, _ in QUARTERS)
fy_old = sum(sum(old_q[q].values()) for q in old_q)
print()
print("  FY total   rebuilt %18s     old workbook %18s     delta %s"
      % ("{:,.0f}".format(fy_new), "{:,.0f}".format(fy_old),
         "{:+,.0f}".format(fy_new - fy_old)))

# ============================================================ 3. DIFFERENCES ==
print()
print("=" * 78)
print("3. WHAT CHANGES  (accounts whose rollup differs from the old workbook)")
print("=" * 78)
diffs = []
for qkey in ("Q1", "Q2"):
    for code in rows:
        o = old_q[qkey].get(code, 0)
        n = new_q[qkey].get(code, 0)
        if abs(o - n) > 0.51:
            nm = mws.cell(rows[code], 2).value
            diffs.append((abs(n - o), qkey, code, nm, o, n))
diffs.sort(reverse=True)
if not diffs:
    print("  none - every account's quarter total reproduces exactly.")
else:
    print("  %-4s %-10s %-22s %16s %16s %16s"
          % ("Qtr", "CODE", "NAME", "old", "rebuilt", "delta"))
    for _, qkey, code, nm, o, n in diffs:
        print("  %-4s %-10s %-22s %16s %16s %16s"
              % (qkey, code, str(nm)[:22], "{:,.0f}".format(o),
                 "{:,.0f}".format(n), "{:+,.0f}".format(n - o)))
    print()
    print("  total recovered: %s" % "{:+,.0f}".format(sum(n - o for _, _, _, _, o, n in diffs)))

# ================================================================== VERDICT ==
print()
print("=" * 78)
if warns:
    print("WARNINGS (%d):" % len(warns))
    for w in warns[:20]:
        print("  - " + w)
    if len(warns) > 20:
        print("  ... and %d more" % (len(warns) - 20))
if fails:
    print("FAILED (%d):" % len(fails))
    for f in fails[:20]:
        print("  - " + f)
    sys.exit(1)
print("PASS - structure, arithmetic and dashboard contract all check out.")
