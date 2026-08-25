import openpyxl

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["FY-(26-27)Q1"]

segments = {}
for r in range(4, ws.max_row + 1):
    code = ws.cell(r, 2).value
    if not code:
        continue
    name = ws.cell(r, 3).value
    segment = ws.cell(r, 6).value
    
    segments.setdefault(segment, []).append((r, code, name))

for seg, accs in segments.items():
    print(f"Segment: {seg}")
    for r, code, name in accs:
        print(f"  Row {r:>2}: {code} - {name}")
