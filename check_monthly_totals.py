import openpyxl

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)

for sheetname in ["APR-2026", "MAY-2026", "JUNE-2026"]:
    ws = wb[sheetname]
    print(f"\n--- {sheetname} ---")
    # Search for rows containing "TOTAL" or formulas at the bottom
    for r in range(ws.max_row - 10, ws.max_row + 1):
        if r < 1:
            continue
        row_vals = [ws.cell(r, c).value for c in range(1, 15)]
        if any(val is not None for val in row_vals):
            print(f"Row {r:>2}: {row_vals}")
