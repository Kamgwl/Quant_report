import openpyxl
import re

wb = openpyxl.load_workbook("Quant Strategy(2026-27).xlsx", data_only=True)
ws = wb["FY-(26-27)Q1"]

COL_CODE     = 2
COL_NAME     = 3
COL_STRATEGY = 5
COL_FUND     = 7
COL_APR      = 8
COL_MAY      = 10
COL_JUN      = 12

print(f"{'Row':<5} | {'Code':<10} | {'Name':<25} | {'Strategy':<15} | {'Apr':<12} | {'May':<12} | {'Jun':<12} | Matches Regex?")
print("-" * 110)

regex_old = re.compile(r'^[A-Z0-9]{3,10}$')
regex_new = re.compile(r'^[A-Z0-9_]{3,10}$')

for r in range(4, ws.max_row + 1):
    code_val = ws.cell(r, COL_CODE).value
    if code_val is None:
        continue
    code = str(code_val).strip()
    if code.upper() in ("CODE", "", "NONE"):
        continue
    
    name = str(ws.cell(r, COL_NAME).value or "").strip()
    strat = str(ws.cell(r, COL_STRATEGY).value or "").strip()
    apr = ws.cell(r, COL_APR).value
    may = ws.cell(r, COL_MAY).value
    jun = ws.cell(r, COL_JUN).value
    
    m_old = bool(regex_old.match(code.upper()))
    m_new = bool(regex_new.match(code.upper()))
    
    match_str = "Old & New" if m_old else ("New Only" if m_new else "Neither")
    
    print(f"{r:<5} | {code:<10} | {name:<25} | {strat:<15} | {str(apr):<12} | {str(may):<12} | {str(jun):<12} | {match_str}")
